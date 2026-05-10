"""Document extraction pipeline for Researcher document management.

The required parser packages are installed by application startup from
requirements.txt. Imports stay inside parser methods so fresh checkouts can
still import the Flask app before startup dependency bootstrap runs.
"""

from __future__ import annotations

import io
import json
import tempfile
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class DocumentExtractionResult:
    """Normalized output from any document parser."""

    text: str | None
    parser_name: str
    parser_version: str | None = None
    extraction_status: str = "ready"
    extraction_quality: str = "standard"
    page_count: int | None = None
    table_count: int | None = None
    image_count: int | None = None
    formula_count: int | None = None
    chart_count: int | None = None
    audio_duration_seconds: float | None = None
    language: str | None = None
    warnings: list[str] = field(default_factory=list)
    structured_json: dict[str, Any] | list[Any] | None = None

    @property
    def has_text(self) -> bool:
        return bool((self.text or "").strip())


class DocumentExtractionService:
    """Extract text and metadata using the required document stack."""

    lightweight_text_extensions = {
        ".txt",
        ".text",
        ".md",
        ".qmd",
        ".rmd",
        ".html",
        ".htm",
        ".csv",
        ".json",
    }

    def extract(
        self, *, filename: str, raw_bytes: bytes, content_type: str | None = None
    ) -> DocumentExtractionResult:
        """Run the parser cascade and return normalized extraction metadata."""

        extension = Path(filename).suffix.lower()
        warnings: list[str] = []

        lightweight = self._extract_lightweight(filename=filename, raw_bytes=raw_bytes)
        if lightweight and extension in self.lightweight_text_extensions:
            return lightweight

        for parser in (
            self._extract_with_docling,
            self._extract_with_pymupdf4llm,
            self._extract_with_unstructured,
        ):
            result = parser(
                filename=filename, raw_bytes=raw_bytes, content_type=content_type
            )
            if result.has_text:
                return result
            warnings.extend(result.warnings)

        if lightweight and lightweight.has_text:
            return lightweight

        return DocumentExtractionResult(
            text=None,
            parser_name="none",
            extraction_status="unavailable",
            extraction_quality="none",
            warnings=warnings
            or ["No parser could extract indexable text from this document."],
        )

    def _extract_with_docling(
        self,
        *,
        filename: str,
        raw_bytes: bytes,
        content_type: str | None = None,
    ) -> DocumentExtractionResult:
        suffix = Path(filename).suffix or ".bin"
        try:
            from docling.document_converter import DocumentConverter
        except Exception as exc:
            return self._failed_result("docling", exc)

        try:
            with self._temporary_document(filename, raw_bytes, suffix) as path:
                result = DocumentConverter().convert(path)
                document = result.document
                text = self._safe_call(document, "export_to_markdown")
                structured_json = self._safe_call(document, "export_to_dict")
                if structured_json is None:
                    structured_json = self._safe_json(
                        self._safe_call(document, "export_to_json")
                    )

                return DocumentExtractionResult(
                    text=(text or "").strip() or None,
                    parser_name="docling",
                    parser_version=self._module_version("docling"),
                    extraction_quality="structured",
                    page_count=self._safe_len(getattr(document, "pages", None)),
                    table_count=self._safe_len(getattr(document, "tables", None)),
                    image_count=self._safe_len(getattr(document, "pictures", None)),
                    structured_json=structured_json,
                )
        except Exception as exc:
            return self._failed_result("docling", exc)

    def _extract_with_pymupdf4llm(
        self,
        *,
        filename: str,
        raw_bytes: bytes,
        content_type: str | None = None,
    ) -> DocumentExtractionResult:
        extension = Path(filename).suffix.lower()
        if extension != ".pdf":
            return DocumentExtractionResult(
                text=None,
                parser_name="pymupdf4llm",
                extraction_status="skipped",
                warnings=["PyMuPDF4LLM fallback only handles PDFs in this pipeline."],
            )
        try:
            import pymupdf4llm
        except Exception as exc:
            return self._failed_result("pymupdf4llm", exc)

        try:
            with self._temporary_document(filename, raw_bytes, ".pdf") as path:
                text = pymupdf4llm.to_markdown(path)
                structured_json = None
                to_json = getattr(pymupdf4llm, "to_json", None)
                if callable(to_json):
                    structured_json = self._safe_json(to_json(path))
                return DocumentExtractionResult(
                    text=(text or "").strip() or None,
                    parser_name="pymupdf4llm",
                    parser_version=self._module_version("pymupdf4llm"),
                    extraction_quality="layout",
                    structured_json=structured_json,
                )
        except Exception as exc:
            return self._failed_result("pymupdf4llm", exc)

    def _extract_with_unstructured(
        self,
        *,
        filename: str,
        raw_bytes: bytes,
        content_type: str | None = None,
    ) -> DocumentExtractionResult:
        try:
            from unstructured.partition.auto import partition
        except Exception as exc:
            return self._failed_result("unstructured", exc)

        try:
            suffix = Path(filename).suffix or ".bin"
            with self._temporary_document(filename, raw_bytes, suffix) as path:
                elements = partition(
                    filename=path, strategy="auto", infer_table_structure=True
                )
                text = "\n\n".join(
                    str(element) for element in elements if str(element).strip()
                )
                table_count = sum(
                    1
                    for element in elements
                    if element.__class__.__name__.lower().endswith("table")
                )
                return DocumentExtractionResult(
                    text=text.strip() or None,
                    parser_name="unstructured",
                    parser_version=self._module_version("unstructured"),
                    extraction_quality="partitioned",
                    table_count=table_count,
                    structured_json=[
                        self._element_to_dict(element) for element in elements
                    ],
                )
        except Exception as exc:
            return self._failed_result("unstructured", exc)

    def _extract_lightweight(
        self, *, filename: str, raw_bytes: bytes
    ) -> DocumentExtractionResult | None:
        extension = Path(filename).suffix.lower()
        if extension in self.lightweight_text_extensions:
            text = raw_bytes.decode("utf-8", errors="ignore").strip()
            return DocumentExtractionResult(
                text=text or None,
                parser_name="lightweight",
                extraction_quality="plain_text",
            )
        if extension == ".pdf":
            return self._extract_pdf_lightweight(raw_bytes)
        if extension == ".docx":
            return self._extract_docx_lightweight(raw_bytes)
        if extension == ".xlsx":
            return self._extract_xlsx_lightweight(raw_bytes)
        return None

    def _extract_pdf_lightweight(self, raw_bytes: bytes) -> DocumentExtractionResult:
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(raw_bytes))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            return DocumentExtractionResult(
                text=text.strip() or None,
                parser_name="pypdf",
                parser_version=self._module_version("pypdf"),
                extraction_quality="basic",
                page_count=len(reader.pages),
            )
        except Exception as exc:
            return self._failed_result("pypdf", exc)

    def _extract_docx_lightweight(self, raw_bytes: bytes) -> DocumentExtractionResult:
        try:
            from docx import Document

            document = Document(io.BytesIO(raw_bytes))
            text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return DocumentExtractionResult(
                text=text.strip() or None,
                parser_name="python-docx",
                parser_version=self._module_version("docx"),
                extraction_quality="basic",
            )
        except Exception as exc:
            return self._failed_result("python-docx", exc)

    def _extract_xlsx_lightweight(self, raw_bytes: bytes) -> DocumentExtractionResult:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                io.BytesIO(raw_bytes), read_only=True, data_only=True
            )
            rows: list[str] = []
            for sheet in workbook.worksheets:
                rows.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value) for value in row if value is not None]
                    if values:
                        rows.append(" | ".join(values))
            return DocumentExtractionResult(
                text="\n".join(rows).strip() or None,
                parser_name="openpyxl",
                parser_version=self._module_version("openpyxl"),
                extraction_quality="basic",
            )
        except Exception as exc:
            return self._failed_result("openpyxl", exc)

    @staticmethod
    def apply_result_to_document(document, result: DocumentExtractionResult) -> None:
        """Copy extraction metadata to a ResearcherDocument model."""

        document.text_content = result.text
        document.extraction_status = (
            result.extraction_status if result.has_text else "unavailable"
        )
        document.extraction_quality = result.extraction_quality
        document.parser_name = result.parser_name
        document.parser_version = result.parser_version
        document.page_count = result.page_count
        document.table_count = result.table_count
        document.image_count = result.image_count
        document.formula_count = result.formula_count
        document.chart_count = result.chart_count
        document.audio_duration_seconds = result.audio_duration_seconds
        document.language = result.language
        document.extraction_warnings = (
            "\n".join(result.warnings) if result.warnings else None
        )

    @staticmethod
    def _failed_result(parser_name: str, exc: Exception) -> DocumentExtractionResult:
        msg = f"{parser_name} failed: {exc}"
        if "No module named" in str(exc) or "ModuleNotFoundError" in str(exc):
            msg += " (Install this feature from Admin > Optional Packages)"
        return DocumentExtractionResult(
            text=None,
            parser_name=parser_name,
            extraction_status="failed",
            extraction_quality="none",
            warnings=[msg],
        )

    @staticmethod
    @contextmanager
    def _temporary_document(filename: str, raw_bytes: bytes, suffix: str):
        safe_suffix = suffix if suffix.startswith(".") else f".{suffix}"
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / f"input{safe_suffix}"
            path.write_bytes(raw_bytes)
            yield str(path)

    @staticmethod
    def _safe_call(obj: Any, method_name: str) -> Any:
        method = getattr(obj, method_name, None)
        if not callable(method):
            return None
        try:
            return method()
        except Exception:
            return None

    @staticmethod
    def _safe_json(value: Any) -> dict[str, Any] | list[Any] | None:
        if isinstance(value, (dict, list)):
            return value
        if not isinstance(value, str):
            return None
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, (dict, list)) else None
        except Exception:
            return None

    @staticmethod
    def _safe_len(value: Any) -> int | None:
        try:
            return len(value) if value is not None else None
        except Exception:
            return None

    @staticmethod
    def _element_to_dict(element: Any) -> dict[str, Any]:
        to_dict = getattr(element, "to_dict", None)
        if callable(to_dict):
            try:
                out = to_dict()
                if isinstance(out, dict):
                    return out
            except Exception:
                pass
        return {
            "type": element.__class__.__name__,
            "text": str(element),
        }

    @staticmethod
    def _module_version(module_name: str) -> str | None:
        try:
            from importlib.metadata import version

            return version(module_name)
        except Exception:
            return None


document_extraction_service = DocumentExtractionService()
